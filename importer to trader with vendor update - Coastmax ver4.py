import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
import os

CONFIG_FILE = "config.txt"

# ==============================
# SAFE ACCOUNTING REFERENCE MAP
# ==============================
# Column E will receive NUMERIC CODES ONLY
raw_reference_map = {
    # =====================
    # Core / COGS Freight
    # =====================
    "material, materials": "50000",

    "international freight, freight costs (ocean), freight cost ocean": "51300",
    "delivery": "55100",
    "fuel surcharge": "56000",
    "overweight": "55800",
    "destination fee, destination terminal handling charges": "55900",

    # =====================
    # Insurance / Courier
    # =====================
    "freight insurance": "51000",
    "courier costs (air), courier cost air, courier air": "51200",

    # =====================
    # Customs / Compliance
    # =====================
    "customs clearance & admin, customs clearance and admin": "51400",
    "isf fee, isf fees": "51500",
    "duties, duty, custom duty 7501, customs 7501, customs": "59240",
    "aes fee": "55700",

    # =====================
    # Inland / Logistics
    # =====================
    "drayage": "51600",
    "destination drayage, drayage (destination)": "59120",
    "transload, transload and final delivery": "55600",
    "pre pull, pre-pull": "59230",

    # =====================
    # Exams / Detention / Yard
    # =====================
    "exam, customs exam fee": "59130",
    "detention": "59140",
    "dry run": "59160",
    "storage": "59170",
    "demurrage, destination demurrage": "59180",
    "destination line demurrage": "55300",
    "per diem": "59190",

    # =====================
    # Chassis / Terminal
    # =====================
    "chassis, destination chassis fee": "59150",
    "terminal fee": "59200",
    "pier pass, destination pierpass, destination pier pass": "59110",

    # =====================
    # Handling / Service
    # =====================
    "handling fees, handling fee": "59210",
    "service fees": "53000",

    # =====================
    # Misc / Others
    # =====================
    "others, others_round up": "59000",
    "others_round up": "59100",

    # =====================
    # Warehouse / EXW
    # =====================
    "exwork, ex-work": "59250",
    "warehouse in/out, warehouse in out": "59260",

    # =====================
    # Bond / Commission
    # =====================
    "bond renewal": "51800",
    "commissions paid": "52000",

    # =====================
    # AMS
    # =====================
    "ams": "59220"
}

# 1. Flatten map for partial matching
reference_map = {}
for key_string, code in raw_reference_map.items():
    for key in key_string.split(","):
        reference_map[key.strip().lower()] = code

# 2. CRITICAL FIX: Sort keywords by length (Longest -> Shortest)
# This prevents "freight" (short) from matching inside "freight insurance" (long).
sorted_keywords = sorted(reference_map.keys(), key=len, reverse=True)


# ==============================
# EXCEL UPDATE LOGIC
# ==============================
def update_excel(file_path):
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        # Track stats for the user
        rows_processed = 0
        matches_found = 0

        for row in ws.iter_rows(min_row=2):
            col_c = row[2]   # Vendor (Index 2 = Column C)
            col_e = row[4]   # Account Code (Index 4 = Column E)
            col_g = row[6]   # Description (Index 6 = Column G)
            col_h = row[7]   # Memo (Index 7 = Column H)
            col_j = row[9]   # Extracted Reference (Index 9 = Column J)

            # Skip empty rows to prevent errors/clutter
            if not col_c.value and not col_g.value:
                continue
            
            rows_processed += 1

            # ---- SAFE Vendor Cleanup (non-destructive) ----
            if col_c.value:
                vendor = str(col_c.value).strip().lower()
                if "perfect gateway enterprises ltd" in vendor:
                    col_c.value = "Perfect Gateway"

            # ---- Cost Code Matching (Using SORTED keywords) ----
            matched = False
            if col_g.value:
                text = str(col_g.value).strip().lower()
                
                # Iterate through Longest keywords first
                for keyword in sorted_keywords:
                    if keyword in text:
                        col_e.value = reference_map[keyword]
                        matched = True
                        matches_found += 1
                        break # Stop checking other words once a match is found

            if not matched:
                col_e.value = "99000"  # Unclassified / review

            # ---- Extract Reference from Column H ----
            if col_h.value:
                h_val = str(col_h.value).strip()
                # Safe check using startswith
                if h_val.startswith("GC Aluminum, Inc:"):
                    extracted = h_val.split("GC Aluminum, Inc:")[-1].strip()
                    if extracted:
                        col_j.value = extracted

        # Save Logic
        folder, original = os.path.split(file_path)
        name, ext = os.path.splitext(original)
        new_path = os.path.join(folder, f"{name}_updatedfortrader{ext}")
        wb.save(new_path)

        messagebox.showinfo(
            "Success", 
            f"✅ Update Complete!\n\nRows Processed: {rows_processed}\nCodes Matched: {matches_found}\n\nSaved to:\n{new_path}"
        )

    except Exception as e:
        messagebox.showerror("Error", f"❌ Update failed:\n{e}")

# ==============================
# CONFIG HANDLING
# ==============================
def load_default_path():
    if os.path.isfile(CONFIG_FILE):
        with open(CONFIG_FILE, "r") as f:
            return f.read().strip()
    return r"C:\Users\eric.gao\Downloads\lastest bill.xlsx"

def save_default_path():
    path = entry_file_path.get()
    if not os.path.isfile(path):
        messagebox.showwarning("Warning", "Please select a valid file.")
        return
    with open(CONFIG_FILE, "w") as f:
        f.write(path)
    messagebox.showinfo("Saved", "✅ Default path saved.")

# ==============================
# GUI
# ==============================
def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        entry_file_path.delete(0, tk.END)
        entry_file_path.insert(0, file_path)

def run_update():
    path = entry_file_path.get()
    if not os.path.isfile(path):
        messagebox.showwarning("Warning", "Invalid file path.")
        return
    update_excel(path)

root = tk.Tk()
root.title("Excel Bill Updater (Accounting Safe)")
root.geometry("520x220") # Slightly taller for better spacing
root.resizable(False, False)

tk.Label(root, text="Excel File Path:").pack(pady=(10, 0))
entry_file_path = tk.Entry(root, width=65)
entry_file_path.pack(pady=5)
entry_file_path.insert(0, load_default_path())

tk.Button(root, text="Browse...", command=browse_file).pack()
tk.Button(root, text="Save as default path", command=save_default_path,
          bg="#2196F3", fg="white").pack(pady=(10, 5))
tk.Button(root, text="Run Update", command=run_update,
          bg="#4CAF50", fg="white", height=2, width=20).pack(pady=5)

root.mainloop()