import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta
from dateutil.parser import parse as du_parse
import os
from collections import defaultdict

CONFIG_FILE = "config.txt"

# ----------------- Helpers -----------------
def find_header_indexes(headers):
    """Find required columns by name (case-insensitive, trimmed)."""
    need = {
        "date": None,
        "expense class": None,
        "expense account": None,
        "expense amount": None
    }
    norm_map = { (str(h).strip().lower() if h is not None else ""): i for i,h in enumerate(headers) }
    for key in need.keys():
        if key in norm_map:
            need[key] = norm_map[key]
        else:
            # fallback: substring match
            found = None
            for h,i in norm_map.items():
                if key in h:
                    found = i
                    break
            if found is None:
                raise Exception(f"Required column not found: '{key}'")
            need[key] = found
    return need

def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%m-%d-%Y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except:
            pass
    try:
        return du_parse(s, dayfirst=False, yearfirst=False)
    except:
        return None

# ----------------- Core logic -----------------
def filter_and_return_removed_only(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    idxs = find_header_indexes(headers)
    date_idx = idxs["date"]
    exp_idx  = idxs["expense class"]
    acct_idx = idxs["expense account"]
    amt_idx  = idxs["expense amount"]

    entries = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        vals = list(r)
        if len(vals) < len(headers):
            vals += [None] * (len(headers) - len(vals))
        exp_val = vals[exp_idx]
        exp_key = str(exp_val).strip().lower() if exp_val not in (None,"") else None
        date_parsed = parse_date(vals[date_idx])
        entries.append({
            "values": vals,
            "exp_val": exp_val,
            "exp_key": exp_key,
            "date": date_parsed
        })

    kept = []
    removed_rows = [headers]
    removed_count = 0

    # Group by Expense Class
    groups = defaultdict(list)
    for e in entries:
        if e["exp_key"]:
            groups[e["exp_key"]].append(e)
        else:
            # Empty Expense Class → remove
            removed_rows.append(e["values"])
            removed_count += 1

    for g in groups.values():
        with_date = [x for x in g if x["date"]]
        no_date   = [x for x in g if not x["date"]]

        # remove no-date rows
        for e in no_date:
            removed_rows.append(e["values"])
            removed_count += 1

        # sort dated rows
        with_date.sort(key=lambda x: x["date"])
        last_kept = None
        for e in with_date:
            if last_kept is None:
                kept.append(e)
                last_kept = e["date"]
            else:
                delta = relativedelta(e["date"], last_kept)
                months = delta.years*12 + delta.months
                if months > 18 or (months == 18 and delta.days > 0):
                    kept.append(e)
                    last_kept = e["date"]
                else:
                    removed_rows.append(e["values"])
                    removed_count += 1

    # Deduplicate removed by Expense Class (keep only one)
    seen_classes = set()
    final_removed = [headers]
    for row in removed_rows[1:]:
        exp_text = str(row[exp_idx]) if row[exp_idx] else ""
        exp_key = exp_text.strip().lower()
        if not exp_key or exp_key in seen_classes:
            continue
        seen_classes.add(exp_key)

        row = list(row)
        if len(row) < len(headers):
            row += [None]*(len(headers)-len(row))
        row[acct_idx] = "45000 Service Revenue"
        if "air" in exp_text.lower():
            row[amt_idx] = 100.00
        else:
            row[amt_idx] = 500.00
        final_removed.append(row)

    # Save result
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Removed"
    for r in final_removed:
        out_ws.append(r)

    folder, fname = os.path.split(file_path)
    name, ext = os.path.splitext(fname)
    out_name = f"{name}_removed_only{ext or '.xlsx'}"
    out_path = os.path.join(folder, out_name)
    out_wb.save(out_path)

    return out_path, len(entries), len(final_removed)-1

# ----------------- GUI -----------------
def load_default_path():
    if os.path.isfile(CONFIG_FILE):
        with open(CONFIG_FILE) as f: return f.read().strip()
    return ""

def save_default_path():
    p = entry_file_path.get()
    if not os.path.isfile(p):
        messagebox.showwarning("Warning","Invalid file.")
        return
    with open(CONFIG_FILE,"w") as f: f.write(p)
    messagebox.showinfo("Saved","✅ Default path saved.")

def browse_file():
    p = filedialog.askopenfilename(filetypes=[("Excel files","*.xlsx;*.xls")])
    if p:
        entry_file_path.delete(0, tk.END)
        entry_file_path.insert(0, p)

def run_process():
    p = entry_file_path.get()
    if not os.path.isfile(p):
        messagebox.showwarning("Warning","Select a valid file.")
        return
    try:
        out, total, removed = filter_and_return_removed_only(p)
        messagebox.showinfo("Done",
            f"Removed rows saved to:\n{out}\n\nProcessed: {total}\nRemoved (unique classes): {removed}")
    except Exception as e:
        messagebox.showerror("Error", str(e))

root = tk.Tk()
root.title("Removed Rows Exporter")
root.geometry("720x250")
root.resizable(False,False)

tk.Label(root,text="Excel File Path:").pack(pady=(10,0))
entry_file_path = tk.Entry(root,width=95)
entry_file_path.pack(pady=5)
entry_file_path.insert(0, load_default_path())
tk.Button(root,text="Browse...",command=browse_file).pack()

tk.Button(root,text="Save this as my default path",command=save_default_path,
          bg="#2196F3",fg="white").pack(pady=(10,5))
tk.Button(root,text="Run Filter",command=run_process,
          bg="#4CAF50",fg="white",height=2).pack(pady=5)

root.mainloop()
