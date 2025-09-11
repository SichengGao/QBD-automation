import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta
import os

CONFIG_FILE = "config.txt"

def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    for fmt in ("%m-%d-%Y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            continue
    return None

def filter_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # find headers
    headers = [cell.value for cell in ws[1]]
    try:
        date_idx = headers.index("Date") + 1
        exp_idx = headers.index("Expense Class") + 1
    except ValueError as e:
        raise Exception(f"Missing required column: {e}")

    # collect rows
    kept_rows = [headers]
    last_kept = {}  # expense_class -> last kept date

    for row in ws.iter_rows(min_row=2, values_only=True):
        exp = row[exp_idx - 1]
        date_val = parse_date(row[date_idx - 1])

        # === Rule 1: skip if Expense Class empty ===
        if exp is None or str(exp).strip() == "":
            continue

        # === Rule 2: filtering by 18-month gap ===
        if exp not in last_kept or last_kept[exp] is None or date_val is None:
            kept_rows.append(row)
            last_kept[exp] = date_val
        else:
            rd = relativedelta(date_val, last_kept[exp])
            months_apart = rd.years * 12 + rd.months
            if months_apart > 18 or (months_apart == 18 and rd.days > 0):
                kept_rows.append(row)
                last_kept[exp] = date_val
            # else: skip row (duplicate within 18 months)

    # write filtered workbook
    new_wb = Workbook()
    new_ws = new_wb.active
    for r in kept_rows:
        new_ws.append(r)

    folder, original_file = os.path.split(file_path)
    name, ext = os.path.splitext(original_file)
    new_filename = f"{name}_filtered for service revenue{ext}"
    new_file_path = os.path.join(folder, new_filename)
    new_wb.save(new_file_path)
    return new_file_path

# ---------- GUI ----------
def load_default_path():
    if os.path.isfile(CONFIG_FILE):
        with open(CONFIG_FILE, "r") as f:
            return f.read().strip()
    return r"C:\Users\eric.gao\Downloads\lastest bill.xlsx"

def save_default_path():
    path = entry_file_path.get()
    if not os.path.isfile(path):
        messagebox.showwarning("Warning", "Select a valid file.")
        return
    with open(CONFIG_FILE, "w") as f:
        f.write(path)
    messagebox.showinfo("Saved", f"✅ Default path saved.")

def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        entry_file_path.delete(0, tk.END)
        entry_file_path.insert(0, file_path)

def run_filter():
    path = entry_file_path.get()
    if not os.path.isfile(path):
        messagebox.showwarning("Warning", "Please select a valid file.")
        return
    try:
        out = filter_excel(path)
        messagebox.showinfo("Success", f"✅ Filtered file saved:\n{out}")
    except Exception as e:
        messagebox.showerror("Error", str(e))

root = tk.Tk()
root.title("Expense Class 18-month Filter")
root.geometry("600x200")
root.resizable(False, False)

tk.Label(root, text="Excel File Path:").pack(pady=(12, 0))
entry_file_path = tk.Entry(root, width=70)
entry_file_path.pack(pady=5)
entry_file_path.insert(0, load_default_path())
tk.Button(root, text="Browse...", command=browse_file).pack()

tk.Button(root, text="Save this as my default path", command=save_default_path, bg="#2196F3", fg="white").pack(pady=(10, 5))
tk.Button(root, text="Run Filter", command=run_filter, bg="#4CAF50", fg="white", height=2).pack(pady=5)

root.mainloop()
