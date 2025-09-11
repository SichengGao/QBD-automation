import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta
import os

CONFIG_FILE = "config.txt"

def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    for fmt in ("%m-%d-%Y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            continue
    return None

def filter_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # find headers
    headers = [cell.value for cell in ws[1]]
    try:
        date_idx = headers.index("Date")
        exp_idx = headers.index("Expense Class")
        acct_idx = headers.index("Expense Account")
        amt_idx = headers.index("Expense Amount")
    except ValueError as e:
        raise Exception(f"Missing required column: {e}")

    removed_rows = [headers]
    last_kept = {}  # expense_class -> last kept date

    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row)
        exp = row[exp_idx]
        date_val = parse_date(row[date_idx])

        # === Rule 1: skip empty Expense Class (goes to removed) ===
        if exp is None or str(exp).strip() == "":
            row[acct_idx] = "45000 Service Revenue"
            if exp and "air" in str(exp).lower():
                row[amt_idx] = "100.00"
            else:
                row[amt_idx] = "500.00"
            removed_rows.append(row)
            continue

        # === Rule 2: filtering by 18-month gap ===
        if exp not in last_kept or last_kept[exp] is None or date_val is None:
            last_kept[exp] = date_val
        else:
            rd = relativedelta(date_val, last_kept[exp])
            months_apart = rd.years * 12 + rd.months
            if months_apart > 18 or (months_apart == 18 and rd.days > 0):
                last_kept[exp] = date_val
            else:
                row[acct_idx] = "45000 Service Revenue"
                if "air" in str(exp).lower():
                    row[amt_idx] = "100.00"
                else:
                    row[amt_idx] = "500.00"
                removed_rows.append(row)

    # write only the removed rows workbook
    new_wb = Workbook()
    ws1 = new_wb.active
    ws1.title = "Removed"
    for r in removed_rows:
        ws1.append(r)

    folder, original_file = os.path.split(file_path)
    name, ext = os.path.splitext(original_file)
    new_filename = f"{name}_removed_only{ext}"
    new_file_path = os.path.join(folder, new_filename)
    new_wb.save(new_file_path)
    return new_file_path

# ---------- GUI ----------
def load_default_path():
    if os.path.isfile(CONFIG_FILE):
        with open(CONFIG_FILE, "r") as f:
            return f.read().strip()
    return r"C:\Users\eric.gao\Downloads\lastest bill.xlsx"

def save_default_path():
    path = entry_file_path.get()
    if not os.path.isfile(path):
        messagebox.showwarning("Warning", "Select a valid file.")
        return
    with open(CONFIG_FILE, "w") as f:
        f.write(path)
    messagebox.showinfo("Saved", f"✅ Default path saved.")

def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        entry_file_path.delete(0, tk.END)
        entry_file_path.insert(0, file_path)

def run_filter():
    path = entry_file_path.get()
    if not os.path.isfile(path):
        messagebox.showwarning("Warning", "Please select a valid file.")
        return
    try:
        out = filter_excel(path)
        messagebox.showinfo("Success", f"✅ Removed rows saved:\n{out}")
    except Exception as e:
        messagebox.showerror("Error", str(e))

root = tk.Tk()
root.title("Removed Rows Exporter")
root.geometry("600x220")
root.resizable(False, False)

tk.Label(root, text="Excel File Path:").pack(pady=(12, 0))
entry_file_path = tk.Entry(root, width=70)
entry_file_path.pack(pady=5)
entry_file_path.insert(0, load_default_path())
tk.Button(root, text="Browse...", command=browse_file).pack()

tk.Button(root, text="Save this as my default path", command=save_default_path, bg="#2196F3", fg="white").pack(pady=(10, 5))
tk.Button(root, text="Run Filter", command=run_filter, bg="#4CAF50", fg="white", height=2).pack(pady=5)

root.mainloop()
