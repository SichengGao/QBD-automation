import os
import re
import pdfplumber
from openpyxl import Workbook
from collections import OrderedDict
import tkinter as tk
from tkinter import filedialog, messagebox

def extract_invoice_data(pdf_path):
    data = OrderedDict()
    data["SOURCE FILE"] = os.path.basename(pdf_path)

    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join(page.extract_text() for page in pdf.pages if page.extract_text())
        lines = text.splitlines()

        for i, line in enumerate(lines):
            line = line.strip()

            # Extract invoice number (with or without /A or /B)
            if re.match(r"INVOICE\s+S\d{6}(/[A-Z])?", line):
                match = re.search(r"S\d{6}(?:/[A-Z])?", line)
                if match:
                    data["INVOICE NUMBER"] = match.group()

            if "INVOICE DATE" in line and "INVOICED" not in line:
                data["INVOICE DATE"] = line.split("INVOICE DATE")[-1].strip()

            elif line.startswith("DUE DATE"):
                data["DUE DATE"] = line.replace("DUE DATE", "").strip()

            elif line.startswith("CUSTOMER ID") and "INVOICED" not in line.upper():
                data["CUSTOMER ID"] = "COASTMAX"

            elif line.startswith("SHIPMENT ") and "DETAILS" not in line:
                data["SHIPMENT"] = line.replace("SHIPMENT", "").strip()

            elif line.startswith("TERMS"):
                data["TERMS"] = line.replace("TERMS", "").strip()

            elif line.startswith("CONSOL NUMBER"):
                data["CONSOL NUMBER"] = line.replace("CONSOL NUMBER", "").strip()

            elif "SHIPPER CONSIGNEE" in line and i + 1 < len(lines):
                data["SHIPPER"] = "EAST ASIA ALUMINUM COMPANY LTD"
                data["CONSIGNEE"] = "COASTMAX INTERNATIONAL"

            elif "GOODS DESCRIPTION" in line and i + 1 < len(lines):
                data["GOODS DESCRIPTION"] = lines[i + 1].strip()

            elif "IMPORT CUSTOMS BROKER" in line and i + 1 < len(lines):
                parts = lines[i + 1].split()
                data["IMPORT BROKER"] = " ".join(parts[:3])
                try:
                    data["WEIGHT"] = parts[3] + " " + parts[4]
                    data["VOLUME"] = parts[5] + " " + parts[6]
                    data["CHARGEABLE VOLUME"] = parts[7] + " " + parts[8]
                    data["PACKAGES"] = parts[9] + " " + parts[10]
                except IndexError:
                    pass

            elif "VESSEL / VOYAGE / IMO" in line and i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                parts = next_line.split()
                if len(parts) >= 3:
                    data["HOUSE B/L"] = parts[-1]
                    data["OCEAN BILL OF LADING"] = parts[-2]
                    data["VESSEL / VOYAGE / IMO"] = " ".join(parts[:-2])

            elif "ORIGIN ETD DESTINATION ETA" in line and i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                date_matches = re.findall(r"\d{2}-[A-Za-z]{3}-\d{2}", next_line)
                if len(date_matches) >= 2:
                    try:
                        etd = date_matches[0]
                        eta = date_matches[1]
                        split_on_etd = next_line.split(etd)
                        split_on_eta = split_on_etd[1].split(eta)
                        origin_part = split_on_etd[0].strip()
                        destination_part = split_on_eta[0].strip()

                        data["ORIGIN"] = origin_part
                        data["ETD"] = etd
                        data["DESTINATION"] = destination_part
                        data["ETA"] = eta
                    except Exception as e:
                        print(f"Error parsing ORIGIN/ETD/DESTINATION/ETA: {e}")

            elif "CONTAINERS" in line and i + 1 < len(lines):
                data["CONTAINERS"] = lines[i + 1].strip()

            elif "DESCRIPTION CHARGES IN USD" in line:
                charge_lines = []
                j = i + 1
                while j < len(lines):
                    charge_line = lines[j].strip()
                    if not charge_line or "TOTAL CHARGES" in charge_line.upper():
                        break
                    charge_lines.append(charge_line)
                    j += 1
                if charge_lines:
                    data["CHARGE DESCRIPTION"] = "; ".join(charge_lines)

            elif "TOTAL USD" in line:
                data["TOTAL USD"] = line.split()[-1]

            elif "CHAIN LOGIC LLC" in line and i + 2 < len(lines):
                data["BANK BENEFICIARY"] = "CHAIN LOGIC LLC"
                data["BANK ADDRESS"] = lines[i + 1].strip() + ", " + lines[i + 2].strip()

            elif "ABA" in line and "SWIFT" in line:
                aba_swift = line.strip().split()
                data["ABA"] = aba_swift[1]
                data["SWIFT"] = aba_swift[3]

            elif "Account" in line and i + 2 < len(lines) and "PINNACLE BANK" in lines[i + 1]:
                data["BANK ACCOUNT"] = line.split("Account")[-1].strip()
                data["BANK NAME"] = "PINNACLE BANK"
                data["BANK LOCATION"] = lines[i + 2].strip()

    return data

def write_all_to_excel(data_list, output_path):
    expanded_data_list = []

    for data in data_list:
        charge_desc = data.get("CHARGE DESCRIPTION", "")
        if charge_desc:
            charges = [x.strip() for x in charge_desc.split(";") if x.strip()]
            for charge in charges:
                match = re.match(r"(.+?)\s+([\d,]+\.\d{2})$", charge)
                if match:
                    desc, amount = match.groups()
                else:
                    desc, amount = charge, ""
                new_data = data.copy()
                new_data["CHARGE DESCRIPTION"] = desc
                new_data["CHARGES IN USD"] = amount
                expanded_data_list.append(new_data)
        else:
            expanded_data_list.append(data)

    all_keys = []
    for row in expanded_data_list:
        for key in row:
            if key not in all_keys:
                all_keys.append(key)

    def reorder_keys(keys):
        reordered = []
        for k in keys:
            if k not in ["CHARGE DESCRIPTION", "CHARGES IN USD", "TOTAL USD"]:
                reordered.append(k)
        if "CHARGE DESCRIPTION" in keys:
            reordered.append("CHARGE DESCRIPTION")
        if "CHARGES IN USD" in keys:
            reordered.append("CHARGES IN USD")
        if "TOTAL USD" in keys:
            reordered.append("TOTAL USD")
        return reordered

    master_keys = reorder_keys(all_keys)

    wb = Workbook()
    ws = wb.active

    for col, key in enumerate(master_keys, start=1):
        ws.cell(row=1, column=col, value=key)

    for row_idx, row_data in enumerate(expanded_data_list, start=2):
        for col_idx, key in enumerate(master_keys, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ""))

    wb.save(output_path)

def run_extraction():
    file_paths = filedialog.askopenfilenames(
        title="Select One or More PDF Invoices",
        filetypes=[("PDF files", "*.pdf")]
    )

    if not file_paths:
        return

    data_list = []
    for path in file_paths:
        try:
            data = extract_invoice_data(path)
            data_list.append(data)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to process {os.path.basename(path)}:\n{str(e)}")
            return

    output_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Save Excel Output As"
    )

    if output_path:
        write_all_to_excel(data_list, output_path)
        messagebox.showinfo("Success", f"✅ Data written to:\n{output_path}")

# GUI
root = tk.Tk()
root.title("PDF Invoice Extractor")
root.geometry("400x200")

label = tk.Label(root, text="Select PDF invoices to extract into Excel", font=("Arial", 12))
label.pack(pady=30)

btn = tk.Button(root, text="Choose PDF File(s)", command=run_extraction, font=("Arial", 12), bg="#4CAF50", fg="white")
btn.pack()

root.mainloop()
